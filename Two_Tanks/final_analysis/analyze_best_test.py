from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import prepare
import test as test_eval

IEEE_SINGLE_COLUMN_WIDTH_IN = 3.5


def read_results(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []

    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        return [dict(row) for row in reader]


def best_validation_row(rows: list[dict[str, str]]) -> dict[str, str] | None:
    best_row: dict[str, str] | None = None
    best_value = math.inf

    for row in rows:
        raw_value = str(row.get("val_RMSE", "")).strip()
        try:
            value = float(raw_value)
        except ValueError:
            continue
        if not math.isfinite(value):
            continue
        if value < best_value:
            best_value = value
            best_row = row

    return best_row


def short_label(description: str) -> str:
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9_-]*", description)
    if not tokens:
        return "best"
    return " ".join(tokens[:2])


def resolve_checkpoint_root(repo_root: Path, checkpoint_set: str) -> Path:
    checkpoint_root = repo_root / "checkpoints"
    if checkpoint_set == "best_so_far":
        return checkpoint_root / "best_so_far"
    return checkpoint_root


def evaluate_checkpoint_set(
    repo_root: Path,
    checkpoint_set: str,
) -> tuple[dict[str, object], np.ndarray, np.ndarray, np.ndarray]:
    _, test_sequence, general_config = prepare.load_datasets_and_config()
    checkpoint_root = (repo_root / general_config["checkpoint_path"]).resolve()
    checkpoint_paths = test_eval.load_checkpoint_paths(checkpoint_root, checkpoint_set)
    checkpoint_paths = [repo_root / p if not p.is_absolute() else p for p in checkpoint_paths]
    if not checkpoint_paths:
        raise FileNotFoundError(f"No fold checkpoints found for checkpoint set {checkpoint_set!r}.")

    fold_predictions = [
        test_eval.predict_with_checkpoint(path, test_sequence=test_sequence, general_config=general_config)
        for path in checkpoint_paths
    ]
    ensemble_prediction = np.mean(np.stack(fold_predictions, axis=0), axis=0)
    y_true = test_sequence.y[0].detach().cpu().numpy()

    metrics = {
        "checkpoint_set": checkpoint_set,
        "checkpoint_root": str(resolve_checkpoint_root(repo_root, checkpoint_set)),
        "checkpoint_paths": [str(path) for path in checkpoint_paths],
        "num_models": len(checkpoint_paths),
    }
    return metrics, y_true, ensemble_prediction, np.stack(fold_predictions, axis=0)


def load_benchmark_test_results(workbook_path: Path, sheet_name: str = "Cascaded Tanks") -> list[float]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required only when you want to read the benchmark workbook. "
            "Install it or run the script without the workbook."
        ) from exc

    if not workbook_path.exists():
        raise FileNotFoundError(f"Missing benchmark workbook: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found in {workbook_path.name}")

    worksheet = workbook[sheet_name]
    values: list[float] = []
    for row_index in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row_index, 8).value
        if cell_value is None:
            continue
        try:
            value = float(cell_value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(value):
            values.append(value)

    if not values:
        raise ValueError(f"No numeric test results found in sheet {sheet_name!r}")

    return values


def rmse(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    return float(np.sqrt(np.mean((y_true - y_pred) ** 2)))


def plot_test_trajectories(
    time_axis: np.ndarray,
    y_true: np.ndarray,
    ensemble_prediction: np.ndarray,
    warmup: int,
    title: str,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fig, ax = plt.subplots(
        figsize=(90 / 25.4, 72 / 25.4),
        constrained_layout=True,
    )
    fig.patch.set_facecolor("#f6f7fb")
    ax.set_facecolor("#ffffff")

    ax.plot(
        time_axis,
        y_true[:, 0],
        linewidth=2.2,
        color="#1d3557",
        label="true output",
        zorder=3,
    )
    ax.plot(
        time_axis,
        ensemble_prediction[:, 0],
        linewidth=2.2,
        color="#e76f51",
        label="estimated output",
        zorder=4,
    )
    ax.axvline(
        time_axis[min(warmup, len(time_axis) - 1)],
        color="#222222",
        linestyle="--",
        linewidth=1.0,
        label="RMSE warmup",
        zorder=2,
    )

    ax.set_title(title, fontsize=8.5, pad=8)
    ax.set_xlabel("Time [s]", fontsize=10)
    ax.set_ylabel(prepare.output_names[0], fontsize=10)
    ax.grid(True, alpha=0.22, linewidth=0.8)
    ax.legend(loc="best", frameon=True, framealpha=0.95, fontsize=8)
    ax.tick_params(labelsize=9)

    ymin = float(np.min(np.concatenate([y_true[:, 0], ensemble_prediction[:, 0]])))
    ymax = float(np.max(np.concatenate([y_true[:, 0], ensemble_prediction[:, 0]])))
    margin = max(0.01, 0.05 * (ymax - ymin if ymax > ymin else 1.0))
    ax.set_ylim(ymin - margin, ymax + margin)

    fig.savefig(output_path, dpi=400)
    plt.close(fig)


def plot_benchmark_distribution(
    benchmark_values: list[float],
    my_test_rmse: float,
    boxplot_output_path: Path,
    histogram_output_path: Path,
) -> None:
    boxplot_output_path.parent.mkdir(parents=True, exist_ok=True)
    histogram_output_path.parent.mkdir(parents=True, exist_ok=True)

    values = np.asarray(benchmark_values, dtype=np.float32)

    fig_box, ax_box = plt.subplots(
        figsize=(90 / 25.4, 72 / 25.4),
        constrained_layout=True,
    )
    fig_box.patch.set_facecolor("#f6f7fb")
    ax_box.set_facecolor("#ffffff")
    ax_box.grid(True, alpha=0.18, linewidth=0.8)
    ax_box.tick_params(labelsize=9)

    ax_box.boxplot(
        values,
        vert=True,
        widths=0.5,
        patch_artist=True,
        boxprops=dict(facecolor="#dbe9f6", color="#4f6d8a", linewidth=1.4),
        whiskerprops=dict(color="#4f6d8a", linewidth=1.2),
        capprops=dict(color="#4f6d8a", linewidth=1.2),
        medianprops=dict(color="#b23a48", linewidth=1.6),
        flierprops=dict(marker="o", markerfacecolor="#9d0208", markeredgecolor="white", markersize=4, alpha=0.55),
    )
    ax_box.scatter(
        [1],
        [my_test_rmse],
        s=80,
        color="#d62728",
        edgecolors="white",
        linewidths=0.9,
        zorder=4,
        label="achieved RMSE",
    )
    ax_box.set_xticks([1])
    ax_box.set_xticklabels([""])
    ax_box.set_ylabel("Test RMSE", fontsize=10)
    ax_box.legend(loc="upper right", frameon=True, framealpha=0.95, fontsize=8)
    fig_box.savefig(boxplot_output_path, dpi=400, bbox_inches="tight")
    plt.close(fig_box)

    fig_hist, ax_hist = plt.subplots(
        figsize=(90 / 25.4, 72 / 25.4),
        constrained_layout=True,
    )
    fig_hist.patch.set_facecolor("#f6f7fb")
    ax_hist.set_facecolor("#ffffff")
    ax_hist.grid(True, alpha=0.18, linewidth=0.8)
    ax_hist.tick_params(labelsize=9)

    ax_hist.hist(
        values,
        bins=25,
        color="#9ecae1",
        edgecolor="#4f6d8a",
        alpha=0.85,
        density=False,
    )
    ax_hist.axvline(
        my_test_rmse,
        color="#d62728",
        linestyle="--",
        linewidth=2.0,
        label="achieved RMSE",
        zorder=5,
    )
    ax_hist.scatter(
        [my_test_rmse],
        [0],
        s=55,
        color="#d62728",
        edgecolors="white",
        linewidths=0.9,
        zorder=6,
    )
    ax_hist.set_xlabel("Test RMSE", fontsize=10)
    ax_hist.set_ylabel("Count", fontsize=10)
    ax_hist.legend(loc="upper right", frameon=True, framealpha=0.95, fontsize=8)
    fig_hist.savefig(histogram_output_path, dpi=400, bbox_inches="tight")
    plt.close(fig_hist)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze the best validation configuration on the saved test ensemble."
    )
    parser.add_argument(
        "--warmup",
        type=int,
        default=None,
        help="Number of initial test outputs to exclude from the RMSE calculation. Defaults to the loaded test sequence.",
    )
    parser.add_argument(
        "--checkpoint-set",
        choices=("current", "best_so_far"),
        default="best_so_far",
        help="Which saved fold checkpoints to analyze.",
    )
    parser.add_argument(
        "--results",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "results.tsv",
        help="Path to the tab-separated experiment log.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().with_name("best_test_trajectory.png"),
        help="PNG path for the test trajectory plot.",
    )
    parser.add_argument(
        "--metrics-output",
        type=Path,
        default=Path(__file__).resolve().with_name("best_test_metrics.json"),
        help="JSON path for the computed test metrics.",
    )
    parser.add_argument(
        "--benchmark-workbook",
        type=Path,
        default=Path(__file__).resolve().with_name("Benchmark Results.xlsx"),
        help="Workbook containing benchmark test results.",
    )
    parser.add_argument(
        "--boxplot-output",
        type=Path,
        default=Path(__file__).resolve().with_name("benchmark_test_boxplot.png"),
        help="PNG path for the benchmark RMSE boxplot.",
    )
    parser.add_argument(
        "--histogram-output",
        type=Path,
        default=Path(__file__).resolve().with_name("benchmark_test_histogram.png"),
        help="PNG path for the benchmark RMSE histogram.",
    )
    parser.add_argument(
        "--skip-benchmark",
        action="store_true",
        help="Skip the benchmark workbook comparison even if the workbook is present.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]

    rows = read_results(args.results)
    best_row = best_validation_row(rows)
    best_description = str(best_row.get("description", "")).strip() if best_row else "best validation"
    best_val_rmse = float(best_row["val_RMSE"]) if best_row and best_row.get("val_RMSE") else math.nan
    best_label = short_label(best_description)

    base_metrics, y_true, ensemble_prediction, fold_predictions = evaluate_checkpoint_set(
        repo_root=repo_root,
        checkpoint_set=args.checkpoint_set,
    )
    _, test_sequence, _ = prepare.load_datasets_and_config()
    time_axis = np.arange(test_sequence.num_samples, dtype=np.float32) * test_sequence.sampling_time

    warmup = test_sequence.warmup if args.warmup is None else int(args.warmup)
    if warmup < 0:
        raise ValueError("warmup must be non-negative")
    if warmup >= len(time_axis):
        raise ValueError("warmup is too large for the available test trajectory")

    test_rmse = rmse(y_true[warmup:], ensemble_prediction[warmup:])
    fold_rmses = [rmse(y_true[warmup:], prediction[warmup:]) for prediction in fold_predictions]

    title = (
        f"Test trajectory for best validation run: {best_label}"
        f" | val RMSE {best_val_rmse:.6f} | test RMSE {test_rmse:.6f}"
    )

    plot_test_trajectories(
        time_axis=time_axis,
        y_true=y_true,
        ensemble_prediction=ensemble_prediction,
        warmup=warmup,
        title=title,
        output_path=args.output,
    )

    benchmark_values: list[float] | None = None
    if not args.skip_benchmark and args.benchmark_workbook.exists():
        benchmark_values = load_benchmark_test_results(args.benchmark_workbook)
        plot_benchmark_distribution(
            benchmark_values=benchmark_values,
            my_test_rmse=test_rmse,
            boxplot_output_path=args.boxplot_output,
            histogram_output_path=args.histogram_output,
        )

    metrics_payload: dict[str, object] = {
        "best_validation_description": best_description,
        "best_validation_label": best_label,
        "best_validation_rmse": best_val_rmse,
        "checkpoint_set_analyzed": args.checkpoint_set,
        "warmup_used_for_test_rmse": warmup,
        "test_rmse_denormalized_after_warmup": test_rmse,
        "fold_test_rmse_denormalized_after_warmup": fold_rmses,
        "source_checkpoint_root": str(resolve_checkpoint_root(repo_root, args.checkpoint_set)),
        "benchmark_workbook": str(args.benchmark_workbook),
        "benchmark_distribution_generated": benchmark_values is not None,
        "benchmark_boxplot_output": str(args.boxplot_output),
        "benchmark_histogram_output": str(args.histogram_output),
        "notes": "The saved test ensemble predictions were re-evaluated after excluding the configured warmup samples.",
        "recomputed_from_checkpoints": base_metrics,
    }
    if benchmark_values is not None:
        metrics_payload["benchmark_results_count"] = len(benchmark_values)
        metrics_payload["benchmark_results_min"] = float(np.min(benchmark_values))
        metrics_payload["benchmark_results_max"] = float(np.max(benchmark_values))

    args.metrics_output.parent.mkdir(parents=True, exist_ok=True)
    args.metrics_output.write_text(json.dumps(metrics_payload, indent=2), encoding="utf-8")

    print(f"Best validation: {best_label} ({best_val_rmse:.6f})")
    print(f"Test RMSE @ warmup={warmup}: {test_rmse:.6f}")
    print(f"Plot saved to: {args.output}")
    if benchmark_values is not None:
        print(f"Boxplot saved to: {args.boxplot_output}")
        print(f"Histogram saved to: {args.histogram_output}")
    else:
        print("Benchmark workbook not found or skipped: benchmark plots not generated.")


if __name__ == "__main__":
    main()
